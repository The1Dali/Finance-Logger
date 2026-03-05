import sys
import csv
import os
from datetime import datetime

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

FINANCE_FILE = os.environ.get("FINANCE_FILE", "finance.csv")
CONFIG_FILE  = os.environ.get("CONFIG_FILE",  "config.csv")
LOG_FILE     = os.environ.get("LOG_FILE",     "log.csv")
EXPORT_FILE  = os.environ.get("EXPORT_FILE",  "finance.xlsx")
GSHEET_ID    = os.environ.get("GSHEET_ID",    "")
_BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
GSHEET_CREDS = os.environ.get("GSHEET_CREDS", os.path.join(_BASE_DIR, "credentials.json"))

FIELDS = ["type", "description", "amount", "date", "notes"]

EXIT_OK            = 0
EXIT_BAD_ARGS      = 1
EXIT_INVALID_INPUT = 2
EXIT_NOT_FOUND     = 3
EXIT_FILE_ERROR    = 4


def print_help():
    print(
        "\nUsage: python finance.py <operation>\n"
        "\nOperations:\n"
        "  income <description> <amount> [notes]     Add an income entry.\n"
        "  expense <description> <amount> [notes]    Add an expense entry.\n"
        "  remove income <id>                        Remove an income entry by ID.\n"
        "  remove expense <id>                       Remove an expense entry by ID.\n"
        "  balance <amount>                          Set the opening balance.\n"
        "  list                                      Print all sheets to the terminal.\n"
        "  view [summary|income|expenses|log]        View a specific sheet (default: all).\n"
        "  export [spreadsheet_id]                   Export to .xlsx and optionally push to Google Sheets.\n"
        "  import [spreadsheet_id]                   Pull new entries from Google Sheets.\n"
        "  clear                                     Wipe all entries.\n"
        "  --help, -h                                Show this help message.\n"
        "\nGoogle Sheets:\n"
        "  Requires a service account credentials JSON (credentials.json next to finance.py).\n"
        "  Import sheet must have a header row: type, description, amount, date, notes\n"
        "  New rows are merged in -- existing entries are never duplicated.\n"
        "\nEnvironment Variables:\n"
        f"  FINANCE_FILE   Path to the finance CSV      (default: {FINANCE_FILE})\n"
        f"  CONFIG_FILE    Path to the config CSV       (default: {CONFIG_FILE})\n"
        f"  LOG_FILE       Path to the log CSV          (default: {LOG_FILE})\n"
        f"  EXPORT_FILE    Path for the .xlsx           (default: {EXPORT_FILE})\n"
        f"  GSHEET_ID      Google Spreadsheet ID        (default: {GSHEET_ID or 'not set'})\n"
        f"  GSHEET_CREDS   Path to service account JSON (default: {GSHEET_CREDS})\n"
        "\nExit Codes:\n"
        "  0  Success\n"
        "  1  Bad or missing arguments\n"
        "  2  Invalid input value\n"
        "  3  Entry not found\n"
        "  4  File error\n"
    )


def main():
    rows = load_finance()

    try:
        match sys.argv[1]:
            case "--help" | "-h":
                print_help()
                sys.exit(EXIT_OK)

            case "income" if len(sys.argv) > 3:
                notes = sys.argv[4] if len(sys.argv) > 4 else ""
                add_entry("INCOME", sys.argv[2], sys.argv[3], rows, notes=notes)

            case "income":
                print("Usage: python finance.py income <description> <amount> [notes]", file=sys.stderr)
                sys.exit(EXIT_BAD_ARGS)

            case "expense" if len(sys.argv) > 3:
                notes = sys.argv[4] if len(sys.argv) > 4 else ""
                add_entry("EXPENSE", sys.argv[2], sys.argv[3], rows, notes=notes)

            case "expense":
                print("Usage: python finance.py expense <description> <amount> [notes]", file=sys.stderr)
                sys.exit(EXIT_BAD_ARGS)

            case "remove" if len(sys.argv) > 3:
                remove_entry(sys.argv[2], sys.argv[3], rows)

            case "remove":
                print("Usage: python finance.py remove <income|expense> <id>", file=sys.stderr)
                sys.exit(EXIT_BAD_ARGS)

            case "balance" if len(sys.argv) > 2:
                set_balance(sys.argv[2])

            case "balance":
                print("Usage: python finance.py balance <amount>", file=sys.stderr)
                sys.exit(EXIT_BAD_ARGS)

            case "list":
                view_sheet(rows, sheet=None)

            case "view":
                sheet = sys.argv[2].lower() if len(sys.argv) > 2 else None
                valid = {"summary", "income", "expenses", "log"}
                if sheet and sheet not in valid:
                    print(f"Unknown sheet '{sheet}'. Choose: summary, income, expenses, log", file=sys.stderr)
                    sys.exit(EXIT_BAD_ARGS)
                view_sheet(rows, sheet=sheet)

            case "export":
                sheet_id = sys.argv[2] if len(sys.argv) > 2 else (GSHEET_ID or None)
                export_xlsx(rows, push_to_gsheet=sheet_id is not None, gsheet_id=sheet_id)

            case "import":
                sheet_id = sys.argv[2] if len(sys.argv) > 2 else GSHEET_ID
                if not sheet_id:
                    print("Error: provide a spreadsheet ID or set the GSHEET_ID environment variable.", file=sys.stderr)
                    sys.exit(EXIT_BAD_ARGS)
                import_from_gsheet(sheet_id, rows)

            case "clear":
                clear_all()

            case _:
                print("Unknown operation. Run 'python finance.py --help' for usage.", file=sys.stderr)
                sys.exit(EXIT_BAD_ARGS)

    except IndexError:
        print("No operation specified. Run 'python finance.py --help' for usage.", file=sys.stderr)
        sys.exit(EXIT_BAD_ARGS)

    sys.exit(EXIT_OK)


def load_finance():
    try:
        with open(FINANCE_FILE, "r", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            for r in rows:
                if "notes" not in r:
                    r["notes"] = ""
            return rows
    except FileNotFoundError:
        try:
            with open(FINANCE_FILE, "w", newline="") as f:
                csv.DictWriter(f, fieldnames=FIELDS).writeheader()
        except OSError as e:
            print(f"Error: could not create '{FINANCE_FILE}': {e}", file=sys.stderr)
            sys.exit(EXIT_FILE_ERROR)
        return []


def save_finance(rows):
    try:
        with open(FINANCE_FILE, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDS)
            writer.writeheader()
            for r in rows:
                writer.writerow({k: r.get(k, "") for k in FIELDS})
    except OSError as e:
        print(f"Error: could not write to '{FINANCE_FILE}': {e}", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)


def load_balance():
    try:
        with open(CONFIG_FILE, "r", newline="") as f:
            for row in csv.DictReader(f):
                if row.get("key") == "opening_balance":
                    return float(row["value"])
    except (FileNotFoundError, ValueError):
        pass
    return 0.0


def save_balance(amount):
    try:
        with open(CONFIG_FILE, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["key", "value"])
            writer.writeheader()
            writer.writerow({"key": "opening_balance", "value": str(amount)})
    except OSError as e:
        print(f"Error: could not write to '{CONFIG_FILE}': {e}", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)


def log_event(action, entry_type="", description="", amount=""):
    has_header = False
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE, "r", encoding="utf-8-sig") as f:
                first = f.readline().strip().lower()
            has_header = "date" in first and "action" in first
        except OSError:
            pass
    date_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    try:
        with open(LOG_FILE, "a", newline="") as f:
            writer = csv.writer(f)
            if not has_header:
                writer.writerow(["date", "action", "type", "description", "amount"])
            writer.writerow([date_str, action, entry_type, description, amount])
    except OSError as e:
        print(f"Warning: could not write to log file '{LOG_FILE}': {e}", file=sys.stderr)


def add_entry(entry_type, description, amount_str, rows, notes="", date_str=None):
    try:
        amount = float(amount_str)
    except ValueError:
        print(f"Error: '{amount_str}' is not a valid number.", file=sys.stderr)
        sys.exit(EXIT_INVALID_INPUT)

    if date_str is None:
        date_str = datetime.now().strftime("%d/%m/%Y")

    rows.append({
        "type":        entry_type,
        "description": description,
        "amount":      str(amount),
        "date":        date_str,
        "notes":       notes,
    })
    save_finance(rows)
    log_event("ADD", entry_type, description, amount)
    note_suffix = f"  [{notes}]" if notes else ""
    print(f"Added {entry_type.lower()} '{description}': {amount:.3f} TND{note_suffix}")


def remove_entry(entry_type, id_str, rows):
    entry_type = entry_type.upper()
    if entry_type not in ("INCOME", "EXPENSE"):
        print(f"Error: type must be 'income' or 'expense', got '{entry_type}'.", file=sys.stderr)
        sys.exit(EXIT_BAD_ARGS)

    try:
        target_id = int(id_str)
    except ValueError:
        print(f"Error: '{id_str}' is not a valid ID.", file=sys.stderr)
        sys.exit(EXIT_INVALID_INPUT)

    typed_indices = [i for i, r in enumerate(rows) if r["type"].upper() == entry_type]
    if target_id < 1 or target_id > len(typed_indices):
        print(f"Error: {entry_type.lower()} entry #{target_id} not found.", file=sys.stderr)
        sys.exit(EXIT_NOT_FOUND)

    actual_index = typed_indices[target_id - 1]
    removed = rows.pop(actual_index)
    save_finance(rows)
    log_event("REMOVE", entry_type, removed["description"], removed["amount"])
    print(f"Removed {entry_type.lower()} #{target_id} '{removed['description']}'.")


def set_balance(amount_str):
    try:
        amount = float(amount_str)
    except ValueError:
        print(f"Error: '{amount_str}' is not a valid number.", file=sys.stderr)
        sys.exit(EXIT_INVALID_INPUT)
    save_balance(amount)
    log_event("BALANCE", "", "Opening balance set", amount)
    print(f"Opening balance set to {amount:.3f} TND")


# ══════════════════════════════════════════════════════════════════════════════
# TERMINAL RENDERING
# ══════════════════════════════════════════════════════════════════════════════

def _ansi(fg=None, bg=None, bold=False, dim=False, italic=False, reset=False):
    if reset:
        return "\033[0m"
    parts = []
    if bold:   parts.append("1")
    if dim:    parts.append("2")
    if italic: parts.append("3")
    if fg:
        r, g, b = int(fg[0:2],16), int(fg[2:4],16), int(fg[4:6],16)
        parts.append(f"38;2;{r};{g};{b}")
    if bg:
        r, g, b = int(bg[0:2],16), int(bg[2:4],16), int(bg[4:6],16)
        parts.append(f"48;2;{r};{g};{b}")
    return f"\033[{';'.join(parts)}m" if parts else ""

R = _ansi(reset=True)

def _cell(text, width, align="left", fg=None, bg=None, bold=False,
          dim=False, italic=False, pad=1):
    text = str(text)
    avail = width - pad * 2
    if len(text) > avail:
        text = text[:avail - 1] + "…"
    if align == "right":
        text = text.rjust(avail)
    elif align == "center":
        text = text.center(avail)
    else:
        text = text.ljust(avail)
    text = " " * pad + text + " " * pad
    style = _ansi(fg=fg, bg=bg, bold=bold, dim=dim, italic=italic)
    return f"{style}{text}{R}"

def _row(*cells):
    print("│" + "│".join(cells) + "│")

def _divider(widths, char="─", left="├", mid="┼", right="┤"):
    segs = [char * w for w in widths]
    print(left + mid.join(segs) + right)

def _top(widths):
    segs = ["─" * w for w in widths]
    print("┌" + "┬".join(segs) + "┐")

def _bottom(widths):
    segs = ["─" * w for w in widths]
    print("└" + "┴".join(segs) + "┘")

def _title_row(text, total_width, fg="FFFFFF", bg="1A56A0"):
    inner = total_width - 2
    styled = _cell(text, inner, align="center", fg=fg, bg=bg, bold=True, pad=2)
    print("│" + styled + "│")

def _section_hdr(text, total_width, fg="FFFFFF", bg="2E75C8"):
    inner = total_width - 2
    styled = _cell(text, inner, align="left", fg=fg, bg=bg, bold=True, pad=2)
    print("│" + styled + "│")

def _spacer_row(total_width, bg="ECF1FB"):
    inner = total_width - 2
    print("│" + _cell("", inner, bg=bg) + "│")

def _tnd(v):
    return f"{float(v):,.3f} TND"


def _sheet_summary(rows):
    opening   = load_balance()
    income    = [r for r in rows if r["type"].upper() == "INCOME"]
    expenses  = [r for r in rows if r["type"].upper() == "EXPENSE"]
    total_in  = sum(float(r["amount"]) for r in income)
    total_exp = sum(float(r["amount"]) for r in expenses)
    gross     = total_in - total_exp
    exp_ratio = (total_exp / total_in * 100) if total_in else 0.0
    net       = opening + gross
    export_dt = datetime.now().strftime("%d/%m/%Y %H:%M")

    W = [36, 22, 24]
    total_w = sum(W) + len(W) + 1

    def s_row(label, value, context, label_fg, label_bg, val_fg, val_bg, ctx_fg, ctx_bg, bold_val=False):
        _row(
            _cell(f"  {label}", W[0], fg=label_fg, bg=label_bg, bold=True),
            _cell(value, W[1], align="right", fg=val_fg, bg=val_bg, bold=bold_val),
            _cell(context, W[2], align="center", fg=ctx_fg, bg=ctx_bg),
        )

    print()
    _top(W)
    _title_row("FINANCIAL SUMMARY", total_w, fg="FFFFFF", bg="1A56A0")
    _divider(W, "─")
    _spacer_row(total_w, bg="ECF1FB")
    _section_hdr("  BALANCE OVERVIEW", total_w, fg="FFFFFF", bg="2E75C8")
    s_row("Opening Balance", _tnd(opening), "Starting funds",
          "0D2B5E","D6E4F7","0D2B5E","D6E4F7","0D2B5E","D6E4F7")
    inc_lbl = f"{len(income)} {'entry' if len(income)==1 else 'entries'}"
    s_row("Total Income", _tnd(total_in), inc_lbl,
          "0B3D1E","D4EDDA","0B3D1E","D4EDDA","0B3D1E","D4EDDA")
    exp_lbl = f"{len(expenses)} {'entry' if len(expenses)==1 else 'entries'}"
    s_row("Total Expenses", _tnd(total_exp), exp_lbl,
          "5C0A0A","F8D7DA","5C0A0A","F8D7DA","5C0A0A","F8D7DA")
    _spacer_row(total_w, bg="ECF1FB")
    _section_hdr("  RESULTS", total_w, fg="FFFFFF", bg="2E75C8")
    gross_ctx = "Surplus ▲" if gross >= 0 else "Deficit ▼"
    s_row("Gross Profit",  _tnd(gross), gross_ctx,
          "0D2B5E","D6E4F7","0D2B5E","D6E4F7","0D2B5E","D6E4F7")
    s_row("Expense Ratio", f"{exp_ratio:.1f}%", f"{exp_ratio:.1f}% of income",
          "0D2B5E","D6E4F7","0D2B5E","D6E4F7","0D2B5E","D6E4F7")
    s_row("Total Profit",  _tnd(gross), gross_ctx,
          "0D2B5E","D6E4F7","0D2B5E","D6E4F7","0D2B5E","D6E4F7")
    net_ctx = "Positive ✔" if net >= 0 else "Negative ✖"
    s_row("Net Balance", _tnd(net), net_ctx,
          "FFFFFF","2E75C8","FFFFFF","2E75C8","FFFFFF","2E75C8", bold_val=True)
    _spacer_row(total_w, bg="ECF1FB")
    _section_hdr("  METADATA", total_w, fg="FFFFFF", bg="7A90B8")
    s_row("Total Entries", str(len(rows)), f"{len(income)} income, {len(expenses)} expense",
          "3A4A6B","EEF2FA","3A4A6B","EEF2FA","3A4A6B","EEF2FA")
    s_row("Last Updated", export_dt, "Export timestamp",
          "3A4A6B","EEF2FA","3A4A6B","EEF2FA","3A4A6B","EEF2FA")
    _bottom(W)


def _sheet_data(rows, entry_type, title, title_bg, hdr_bg, id_fg, total_fg, total_bg):
    typed = [r for r in rows if r["type"].upper() == entry_type]
    W     = [5, 30, 18, 13, 22]
    total_w = sum(W) + len(W) + 1

    print()
    _top(W)
    _title_row(title, total_w, fg="FFFFFF", bg=title_bg)
    _divider(W, "─")
    _row(
        _cell("ID",           W[0], align="center", fg="FFFFFF", bg=hdr_bg, bold=True),
        _cell("Description",  W[1], align="left",   fg="FFFFFF", bg=hdr_bg, bold=True),
        _cell("Amount (TND)", W[2], align="right",  fg="FFFFFF", bg=hdr_bg, bold=True),
        _cell("Date",         W[3], align="center", fg="FFFFFF", bg=hdr_bg, bold=True),
        _cell("Notes",        W[4], align="left",   fg="FFFFFF", bg=hdr_bg, bold=True),
    )
    _divider(W, "─")

    if not typed:
        _row(
            _cell("",  W[0]),
            _cell("No entries yet.", W[1], italic=True, dim=True),
            _cell("",  W[2]), _cell("",  W[3]), _cell("",  W[4]),
        )
    else:
        total = 0.0
        for i, r in enumerate(typed):
            alt = i % 2 == 1
            bg  = "EEF3FC" if alt else "FFFFFF"
            amt = float(r["amount"])
            total += amt
            _row(
                _cell(str(i+1),          W[0], align="center", fg=id_fg,    bg=bg),
                _cell(r["description"],  W[1], align="left",   fg="1A1A2E", bg=bg),
                _cell(_tnd(amt),         W[2], align="right",  fg=total_fg, bg=bg, bold=True),
                _cell(r["date"],         W[3], align="center", fg="1A1A2E", bg=bg),
                _cell(r.get("notes",""), W[4], align="left",   fg="4A4A4A", bg=bg, italic=True),
            )
        _divider(W, "─")
        _row(
            _cell("",          W[0], bg=total_bg),
            _cell("TOTAL",     W[1], align="left",  fg=total_fg, bg=total_bg, bold=True),
            _cell(_tnd(total), W[2], align="right", fg=total_fg, bg=total_bg, bold=True),
            _cell("",          W[3], bg=total_bg),
            _cell("",          W[4], bg=total_bg),
        )
    _bottom(W)


def _sheet_income(rows):
    _sheet_data(rows, "INCOME", "INCOME",
                title_bg="1A7A3A", hdr_bg="28A745",
                id_fg="1A7A3A", total_fg="0B3D1E", total_bg="D4EDDA")

def _sheet_expenses(rows):
    _sheet_data(rows, "EXPENSE", "EXPENSES",
                title_bg="B52525", hdr_bg="DC3545",
                id_fg="B52525", total_fg="5C0A0A", total_bg="F8D7DA")


_LOG_FIELDS = ["date", "action", "type", "description", "amount"]

def _read_log(newest_first=False):
    if not os.path.exists(LOG_FILE):
        return []
    try:
        with open(LOG_FILE, "r", newline="", encoding="utf-8-sig") as f:
            raw = list(csv.reader(f))
        if not raw:
            return []
        first = [c.strip().lower() for c in raw[0]]
        if "date" in first and "action" in first:
            fields, data = first, raw[1:]
        else:
            fields, data = _LOG_FIELDS, raw
        rows = [
            {fields[i]: (row[i] if i < len(row) else "")
             for i in range(len(fields))}
            for row in data if any(c.strip() for c in row)
        ]
        return list(reversed(rows)) if newest_first else rows
    except OSError:
        return []


def _sheet_log():
    ACTION_COLORS = {
        "ADD":     ("155724","D4EDDA"),
        "REMOVE":  ("721C24","F8D7DA"),
        "CLEAR":   ("856404","FFF3CD"),
        "BALANCE": ("0D2B5E","D6E4F7"),
        "IMPORT":  ("4A0E72","E8D5F5"),
        "EXPORT":  ("1A5276","D6EAF8"),
    }
    log_rows = _read_log(newest_first=True)
    W = [19, 10, 9, 30, 18]
    total_w = sum(W) + len(W) + 1

    print()
    _top(W)
    _title_row("ACTIVITY LOG", total_w, fg="FFFFFF", bg="2C2C2C")
    _divider(W, "─")
    _row(
        _cell("Date",        W[0], align="center", fg="FFFFFF", bg="4A4A4A", bold=True),
        _cell("Action",      W[1], align="center", fg="FFFFFF", bg="4A4A4A", bold=True),
        _cell("Type",        W[2], align="center", fg="FFFFFF", bg="4A4A4A", bold=True),
        _cell("Description", W[3], align="left",   fg="FFFFFF", bg="4A4A4A", bold=True),
        _cell("Amount",      W[4], align="right",  fg="FFFFFF", bg="4A4A4A", bold=True),
    )
    _divider(W, "─")

    if not log_rows:
        _row(
            _cell("", W[0]),
            _cell("No log entries yet.", W[1]+W[2]+W[3]+3, italic=True, dim=True),
            _cell("", W[4]),
        )
    else:
        for i, r in enumerate(log_rows):
            alt    = i % 2 == 1
            bg_row = "EEF3FC" if alt else "FFFFFF"
            action = r.get("action","").strip().upper()
            afg, abg = ACTION_COLORS.get(action, ("444444","F5F5F5"))
            amt_str  = r.get("amount","").strip()
            try:
                amt_disp = _tnd(float(amt_str))
            except (ValueError, TypeError):
                amt_disp = amt_str
            _row(
                _cell(r.get("date",""),        W[0], align="center", fg="1A1A2E", bg=bg_row),
                _cell(action,                  W[1], align="center", fg=afg,     bg=abg,    bold=True),
                _cell(r.get("type",""),        W[2], align="center", fg="1A1A2E",bg=bg_row),
                _cell(r.get("description",""), W[3], align="left",   fg="1A1A2E",bg=bg_row),
                _cell(amt_disp,                W[4], align="right",  fg="1A1A2E",bg=bg_row),
            )
    _bottom(W)


def view_sheet(rows, sheet=None):
    if sheet in (None, "summary"):
        _sheet_summary(rows)
    if sheet in (None, "income"):
        _sheet_income(rows)
    if sheet in (None, "expenses"):
        _sheet_expenses(rows)
    if sheet in (None, "log"):
        _sheet_log()
    print()


def clear_all():
    try:
        with open(FINANCE_FILE, "w", newline="") as f:
            csv.DictWriter(f, fieldnames=FIELDS).writeheader()
    except OSError as e:
        print(f"Error: could not clear '{FINANCE_FILE}': {e}", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)
    log_event("CLEAR")
    print("All entries cleared.")


COLUMN_MAP = {
    "type":        ["type", "entry type", "transaction type", "category"],
    "description": ["description", "desc", "item", "name", "details"],
    "amount":      ["amount", "amount (tnd)", "amount (usd)", "value", "price", "total"],
    "date":        ["date", "day", "when"],
    "notes":       ["notes", "note", "comments", "comment", "remarks"],
}

SKIP_TABS       = {"summary", "metadata", "template", "readme", "instructions"}
HEADER_KEYWORDS = {"description", "desc", "amount", "value", "price", "date", "day", "item", "name"}


def _find_col(headers, field):
    hl = [h.strip().lower() for h in headers]
    for name in COLUMN_MAP[field]:
        if name in hl:
            return hl.index(name)
    return None


def _get_cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx]).strip() if row[idx] else ""


def import_from_gsheet(sheet_id, existing_rows):
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError:
        print("Error: Google API libraries are required.\nRun: pip install google-api-python-client google-auth", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)

    if not os.path.exists(GSHEET_CREDS):
        print(f"Error: credentials file '{GSHEET_CREDS}' not found.", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)

    print(f"Connecting to Google Sheets ({sheet_id})...")

    try:
        creds = service_account.Credentials.from_service_account_file(
            GSHEET_CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
        service     = build("sheets", "v4", credentials=creds)
        spreadsheet = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        sheets      = spreadsheet.get("sheets", [])
    except Exception as e:
        print(f"Error: could not connect to Google Sheets: {e}", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)

    if not sheets:
        print("No sheets found in this spreadsheet.")
        return

    print(f"Found {len(sheets)} tab(s): {[s['properties']['title'] for s in sheets]}")
    total_added = total_skipped = total_invalid = 0

    for sheet in sheets:
        tab_name = sheet["properties"]["title"]
        if tab_name.lower() in SKIP_TABS:
            print(f"\nSkipping tab: {tab_name}")
            continue

        print(f"\nProcessing tab: {tab_name}")
        try:
            result   = service.spreadsheets().values().get(
                spreadsheetId=sheet_id, range=f"'{tab_name}'!A:F").execute()
            raw_rows = result.get("values", [])
        except Exception as e:
            print(f"  Warning: could not read tab '{tab_name}': {e}")
            continue

        if len(raw_rows) < 2:
            print("  Tab is empty or header-only -- nothing to import.")
            continue

        headers = None
        data_start_idx = 0
        for i in range(min(3, len(raw_rows))):
            row_words = {cell.strip().lower() for cell in raw_rows[i]}
            if row_words & HEADER_KEYWORDS:
                headers, data_start_idx = raw_rows[i], i + 1
                print(f"  Found headers at row {i + 1}: {headers}")
                break

        if not headers:
            print(f"  Could not find valid headers in first 3 rows: {raw_rows[:3]}")
            continue

        col_idx = {field: _find_col(headers, field) for field in COLUMN_MAP}
        missing = [f for f in ("description", "amount", "date") if col_idx[f] is None]
        if missing:
            print(f"  Skipping -- missing required columns: {', '.join(missing)}")
            continue

        existing_keys = {
            (r["type"].upper(), r["description"].strip(), r["amount"].strip(), r["date"].strip())
            for r in existing_rows
        }
        added = skipped = invalid = 0

        for raw in raw_rows[data_start_idx:]:
            if not any(str(c).strip() for c in raw):
                continue
            entry_type  = _get_cell(raw, col_idx["type"]).upper()
            description = _get_cell(raw, col_idx["description"])
            amount_str  = _get_cell(raw, col_idx["amount"])
            date_str    = _get_cell(raw, col_idx["date"])
            notes       = _get_cell(raw, col_idx["notes"])

            if not entry_type:
                if "income" in tab_name.lower():   entry_type = "INCOME"
                elif "expense" in tab_name.lower(): entry_type = "EXPENSE"
                else:
                    invalid += 1; continue

            if entry_type not in ("INCOME", "EXPENSE"):
                invalid += 1; continue

            try:
                clean  = amount_str.replace("TND","").replace("USD","").replace(",","").replace("$","").strip()
                amount = float(clean)
            except ValueError:
                invalid += 1; continue

            if not date_str:
                date_str = datetime.now().strftime("%d/%m/%Y")

            key = (entry_type, description, str(amount), date_str)
            if key in existing_keys:
                skipped += 1; continue

            existing_rows.append({"type": entry_type, "description": description,
                                   "amount": str(amount), "date": date_str, "notes": notes})
            existing_keys.add(key)
            log_event("IMPORT", entry_type, description, amount)
            added += 1

        print(f"  Tab complete: {added} added, {skipped} already existed, {invalid} invalid.")
        total_added += added; total_skipped += skipped; total_invalid += invalid

    if total_added > 0:
        save_finance(existing_rows)

    print(f"\n{'=' * 50}")
    print(f"Import complete: {total_added} added | {total_skipped} already existed | {total_invalid} invalid")


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE SHEETS FORMATTING
# ══════════════════════════════════════════════════════════════════════════════

def _build_gsheet_format_requests(tab_ids, n_inc, n_exp, log_rows):
    """Build all batchUpdate formatting requests for every tab."""

    def color(h):
        h = h.lstrip("#")
        return {"red": int(h[0:2],16)/255, "green": int(h[2:4],16)/255, "blue": int(h[4:6],16)/255}

    def rc(sid, r1, r2, c1, c2, fmt, fields):
        return {"repeatCell": {
            "range": {"sheetId": sid, "startRowIndex": r1, "endRowIndex": r2,
                      "startColumnIndex": c1, "endColumnIndex": c2},
            "cell": {"userEnteredFormat": fmt},
            "fields": fields,
        }}

    FULL  = "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)"
    BG    = "userEnteredFormat.backgroundColor"

    def merge(sid, r1, r2, c1, c2):
        return {"mergeCells": {
            "range": {"sheetId": sid, "startRowIndex": r1, "endRowIndex": r2,
                      "startColumnIndex": c1, "endColumnIndex": c2},
            "mergeType": "MERGE_ALL",
        }}

    def row_h(sid, r1, r2, px):
        return {"updateDimensionProperties": {
            "range": {"sheetId": sid, "dimension": "ROWS", "startIndex": r1, "endIndex": r2},
            "properties": {"pixelSize": px}, "fields": "pixelSize",
        }}

    def col_w(sid, c1, c2, px):
        return {"updateDimensionProperties": {
            "range": {"sheetId": sid, "dimension": "COLUMNS", "startIndex": c1, "endIndex": c2},
            "properties": {"pixelSize": px}, "fields": "pixelSize",
        }}

    def freeze(sid, n=3):
        return {"updateSheetProperties": {
            "properties": {"sheetId": sid, "gridProperties": {"frozenRowCount": n}},
            "fields": "gridProperties.frozenRowCount",
        }}

    def txt(fg, bold=False, size=10, italic=False):
        t = {"foregroundColor": color(fg), "bold": bold, "fontSize": size}
        if italic: t["italic"] = True
        return t

    def title_fmt(bg, fg="FFFFFF"):
        return {"backgroundColor": color(bg), "textFormat": txt(fg, bold=True, size=14),
                "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE"}

    def hdr_fmt(bg, fg="FFFFFF", align="CENTER"):
        return {"backgroundColor": color(bg), "textFormat": txt(fg, bold=True, size=10),
                "horizontalAlignment": align, "verticalAlignment": "MIDDLE"}

    def cell_fmt(bg, fg="1A1A2E", align="LEFT", bold=False, size=10, italic=False):
        return {"backgroundColor": color(bg), "textFormat": txt(fg, bold=bold, size=size, italic=italic),
                "horizontalAlignment": align, "verticalAlignment": "MIDDLE"}

    reqs = []

    # ── SUMMARY ──────────────────────────────────────────────────────────
    sid = tab_ids["Summary"]
    reqs += [
        col_w(sid, 0, 1, 260), col_w(sid, 1, 2, 175), col_w(sid, 2, 3, 190),
        # Row 1: title banner
        row_h(sid, 0, 1, 44), merge(sid, 0, 1, 0, 3),
        rc(sid, 0, 1, 0, 3, title_fmt("1A56A0"), FULL),
        # Row 2: spacer
        row_h(sid, 1, 2, 8),
        rc(sid, 1, 2, 0, 3, {"backgroundColor": color("ECF1FB")}, BG),
        # Row 3: BALANCE OVERVIEW header
        row_h(sid, 2, 3, 20),
        rc(sid, 2, 3, 0, 3, hdr_fmt("2E75C8", align="LEFT"), FULL),
        # Row 4: Opening Balance
        row_h(sid, 3, 4, 26),
        rc(sid, 3, 4, 0, 1, cell_fmt("D6E4F7","0D2B5E","LEFT", bold=True), FULL),
        rc(sid, 3, 4, 1, 2, cell_fmt("D6E4F7","0D2B5E","RIGHT"), FULL),
        rc(sid, 3, 4, 2, 3, cell_fmt("D6E4F7","0D2B5E","CENTER"), FULL),
        # Row 5: Total Income
        row_h(sid, 4, 5, 26),
        rc(sid, 4, 5, 0, 1, cell_fmt("D4EDDA","0B3D1E","LEFT", bold=True), FULL),
        rc(sid, 4, 5, 1, 2, cell_fmt("D4EDDA","0B3D1E","RIGHT"), FULL),
        rc(sid, 4, 5, 2, 3, cell_fmt("D4EDDA","0B3D1E","CENTER"), FULL),
        # Row 6: Total Expenses
        row_h(sid, 5, 6, 26),
        rc(sid, 5, 6, 0, 1, cell_fmt("F8D7DA","5C0A0A","LEFT", bold=True), FULL),
        rc(sid, 5, 6, 1, 2, cell_fmt("F8D7DA","5C0A0A","RIGHT"), FULL),
        rc(sid, 5, 6, 2, 3, cell_fmt("F8D7DA","5C0A0A","CENTER"), FULL),
        # Row 7: spacer
        row_h(sid, 6, 7, 8),
        rc(sid, 6, 7, 0, 3, {"backgroundColor": color("ECF1FB")}, BG),
        # Row 8: RESULTS header
        row_h(sid, 7, 8, 20),
        rc(sid, 7, 8, 0, 3, hdr_fmt("2E75C8", align="LEFT"), FULL),
    ]
    # Rows 9-11: result data (Gross Profit, Expense Ratio, Total Profit)
    for ri in range(8, 11):
        reqs += [
            row_h(sid, ri, ri+1, 26),
            rc(sid, ri, ri+1, 0, 1, cell_fmt("D6E4F7","0D2B5E","LEFT", bold=True), FULL),
            rc(sid, ri, ri+1, 1, 2, cell_fmt("D6E4F7","0D2B5E","RIGHT"), FULL),
            rc(sid, ri, ri+1, 2, 3, cell_fmt("D6E4F7","0D2B5E","CENTER"), FULL),
        ]
    reqs += [
        # Row 12: Net Balance
        row_h(sid, 11, 12, 26),
        rc(sid, 11, 12, 0, 1, cell_fmt("2E75C8","FFFFFF","LEFT",  bold=True), FULL),
        rc(sid, 11, 12, 1, 2, cell_fmt("2E75C8","FFFFFF","RIGHT", bold=True), FULL),
        rc(sid, 11, 12, 2, 3, cell_fmt("2E75C8","FFFFFF","CENTER",bold=True), FULL),
        # Row 13: spacer
        row_h(sid, 12, 13, 8),
        rc(sid, 12, 13, 0, 3, {"backgroundColor": color("ECF1FB")}, BG),
        # Row 14: METADATA header
        row_h(sid, 13, 14, 20),
        rc(sid, 13, 14, 0, 3, hdr_fmt("7A90B8", align="LEFT"), FULL),
        # Rows 15-16: metadata data
        row_h(sid, 14, 15, 26),
        rc(sid, 14, 15, 0, 1, cell_fmt("EEF2FA","3A4A6B","LEFT", bold=True), FULL),
        rc(sid, 14, 15, 1, 2, cell_fmt("EEF2FA","3A4A6B","RIGHT"), FULL),
        rc(sid, 14, 15, 2, 3, cell_fmt("EEF2FA","3A4A6B","CENTER"), FULL),
        row_h(sid, 15, 16, 26),
        rc(sid, 15, 16, 0, 1, cell_fmt("EEF2FA","3A4A6B","LEFT", bold=True), FULL),
        rc(sid, 15, 16, 1, 2, cell_fmt("EEF2FA","3A4A6B","RIGHT"), FULL),
        rc(sid, 15, 16, 2, 3, cell_fmt("EEF2FA","3A4A6B","CENTER"), FULL),
    ]

    # ── INCOME ───────────────────────────────────────────────────────────
    sid = tab_ids["Income"]
    reqs += [
        col_w(sid, 0, 1,  55), col_w(sid, 1, 2, 260), col_w(sid, 2, 3, 145),
        col_w(sid, 3, 4, 130), col_w(sid, 4, 5, 320),
        row_h(sid, 0, 1, 40), merge(sid, 0, 1, 0, 5),
        rc(sid, 0, 1, 0, 5, title_fmt("1A7A3A"), FULL),
        row_h(sid, 1, 2, 6),
        rc(sid, 1, 2, 0, 5, {"backgroundColor": color("FFFFFF")}, BG),
        row_h(sid, 2, 3, 24),
        rc(sid, 2, 3, 0, 5, hdr_fmt("28A745"), FULL),
        freeze(sid, 3),
    ]
    for i in range(max(n_inc, 1)):
        ri = 3 + i
        bg = "EEF3FC" if i % 2 == 1 else "FFFFFF"
        reqs += [
            row_h(sid, ri, ri+1, 22),
            rc(sid, ri, ri+1, 0, 1, cell_fmt(bg, align="CENTER"), FULL),
            rc(sid, ri, ri+1, 1, 2, cell_fmt(bg, align="LEFT"),   FULL),
            rc(sid, ri, ri+1, 2, 3, cell_fmt(bg, align="RIGHT"),  FULL),
            rc(sid, ri, ri+1, 3, 4, cell_fmt(bg, align="CENTER"), FULL),
            rc(sid, ri, ri+1, 4, 5, cell_fmt(bg, align="LEFT"),   FULL),
        ]

    # ── EXPENSES ─────────────────────────────────────────────────────────
    sid = tab_ids["Expenses"]
    reqs += [
        col_w(sid, 0, 1,  55), col_w(sid, 1, 2, 260), col_w(sid, 2, 3, 145),
        col_w(sid, 3, 4, 130), col_w(sid, 4, 5, 320),
        row_h(sid, 0, 1, 40), merge(sid, 0, 1, 0, 5),
        rc(sid, 0, 1, 0, 5, title_fmt("B52525"), FULL),
        row_h(sid, 1, 2, 6),
        rc(sid, 1, 2, 0, 5, {"backgroundColor": color("FFFFFF")}, BG),
        row_h(sid, 2, 3, 24),
        rc(sid, 2, 3, 0, 5, hdr_fmt("DC3545"), FULL),
        freeze(sid, 3),
    ]
    for i in range(max(n_exp, 1)):
        ri = 3 + i
        bg = "EEF3FC" if i % 2 == 1 else "FFFFFF"
        reqs += [
            row_h(sid, ri, ri+1, 22),
            rc(sid, ri, ri+1, 0, 1, cell_fmt(bg, align="CENTER"), FULL),
            rc(sid, ri, ri+1, 1, 2, cell_fmt(bg, align="LEFT"),   FULL),
            rc(sid, ri, ri+1, 2, 3, cell_fmt(bg, align="RIGHT"),  FULL),
            rc(sid, ri, ri+1, 3, 4, cell_fmt(bg, align="CENTER"), FULL),
            rc(sid, ri, ri+1, 4, 5, cell_fmt(bg, align="LEFT"),   FULL),
        ]

    # ── LOG ──────────────────────────────────────────────────────────────
    sid = tab_ids["Log"]
    ACTION_BG = {
        "ADD":     "D4EDDA", "REMOVE":  "F8D7DA", "CLEAR":   "FFF3CD",
        "BALANCE": "D6E4F7", "IMPORT":  "E8D5F5", "EXPORT":  "D6EAF8",
    }
    ACTION_FG = {
        "ADD":     "155724", "REMOVE":  "721C24", "CLEAR":   "856404",
        "BALANCE": "0D2B5E", "IMPORT":  "4A0E72", "EXPORT":  "1A5276",
    }
    reqs += [
        col_w(sid, 0, 1, 190), col_w(sid, 1, 2, 100), col_w(sid, 2, 3, 90),
        col_w(sid, 3, 4, 290), col_w(sid, 4, 5, 175),
        row_h(sid, 0, 1, 40), merge(sid, 0, 1, 0, 5),
        rc(sid, 0, 1, 0, 5, title_fmt("2C2C2C"), FULL),
        row_h(sid, 1, 2, 6),
        rc(sid, 1, 2, 0, 5, {"backgroundColor": color("FFFFFF")}, BG),
        row_h(sid, 2, 3, 24),
        rc(sid, 2, 3, 0, 5, hdr_fmt("4A4A4A"), FULL),
        freeze(sid, 3),
    ]
    for i, r in enumerate(log_rows):
        ri     = 3 + i
        action = r.get("action","").strip().upper()
        abg    = ACTION_BG.get(action, "F5F5F5")
        afg    = ACTION_FG.get(action, "444444")
        reqs += [
            row_h(sid, ri, ri+1, 22),
            rc(sid, ri, ri+1, 0, 1, cell_fmt("FFFFFF", align="CENTER"), FULL),  # Date
            rc(sid, ri, ri+1, 1, 2,                                               # Action badge
               {"backgroundColor": color(abg),
                "textFormat": txt(afg, bold=True, size=9),
                "horizontalAlignment": "CENTER",
                "verticalAlignment": "MIDDLE"}, FULL),
            rc(sid, ri, ri+1, 2, 3, cell_fmt("FFFFFF", align="CENTER"), FULL),  # Type
            rc(sid, ri, ri+1, 3, 4, cell_fmt("FFFFFF", align="LEFT"),   FULL),  # Description
            rc(sid, ri, ri+1, 4, 5, cell_fmt("FFFFFF", align="RIGHT"),  FULL),  # Amount
        ]

    return reqs


# ══════════════════════════════════════════════════════════════════════════════
# XLSX EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def export_xlsx(rows, push_to_gsheet=False, gsheet_id=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.styles.colors import Color
    except ImportError:
        print("Error: openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)

    FONT        = "Arial"
    WHITE       = "FFFFFF"
    BODY_ALT    = "EEF3FC"
    BORDER_CLR  = "C5D0E8"
    NAVY        = "1B2A6B"
    BLUE_DARK   = "1A56A0"
    BLUE_MID    = "2E75C8"
    BLUE_LIGHT  = "D6E4F7"
    BLUE_FG     = "0D2B5E"
    GREEN_DARK  = "1A7A3A"
    GREEN_MID   = "28A745"
    GREEN_LIGHT = "D4EDDA"
    GREEN_FG    = "0B3D1E"
    RED_DARK    = "B52525"
    RED_MID     = "DC3545"
    RED_LIGHT   = "F8D7DA"
    RED_FG      = "5C0A0A"
    CHARCOAL    = "2C2C2C"
    GREY_DARK   = "4A4A4A"
    META_BG     = "EEF2FA"
    META_FG     = "3A4A6B"
    META_ACC    = "7A90B8"
    TND_FMT     = '#,##0.000 "TND"'
    PCT_FMT     = "0.0%"

    wb = Workbook()

    def _argb(h):
        h = h.lstrip("#")
        return h if len(h) == 8 else "FF" + h

    def solid(h):
        return PatternFill("solid", fgColor=Color(rgb=_argb(h)))

    def _side(style, color):
        return Side(style=style, color=_argb(color))

    def _box(c, w="thin"):
        s = _side(w, c)
        return Border(left=s, right=s, top=s, bottom=s)

    def _left_accent(a, b=BORDER_CLR):
        return Border(left=_side("medium",a), right=_side("thin",b),
                      top=_side("thin",b), bottom=_side("thin",b))

    def _right_accent(a, b=BORDER_CLR):
        return Border(left=_side("thin",b), right=_side("medium",a),
                      top=_side("thin",b), bottom=_side("thin",b))

    def title_cell(ws, merge_range, text, bg, fg=WHITE, row_h=40):
        ws.merge_cells(merge_range)
        first_ref = merge_range.split(":")[0]
        c = ws[first_ref]
        c.value     = text
        c.font      = Font(name=FONT, bold=True, size=14, color=_argb(fg))
        c.fill      = solid(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(bottom=_side("medium", bg))
        first_row   = int("".join(filter(str.isdigit, first_ref)))
        ws.row_dimensions[first_row].height = row_h

    def hdr(cell, text, bg, fg=WHITE, align="center"):
        cell.value     = text
        cell.font      = Font(name=FONT, bold=True, size=10, color=_argb(fg))
        cell.fill      = solid(bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = _box(bg)

    def dat(cell, value, align="left", alt=False, fmt=None, wrap=False):
        cell.value     = value
        cell.font      = Font(name=FONT, size=10, color=_argb("1A1A2E"))
        cell.fill      = solid(BODY_ALT if alt else WHITE)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border    = _box(BORDER_CLR)
        if fmt:
            cell.number_format = fmt

    income_rows  = [r for r in rows if r["type"].upper() == "INCOME"]
    expense_rows = [r for r in rows if r["type"].upper() == "EXPENSE"]
    opening      = load_balance()
    n_inc        = len(income_rows)
    n_exp        = len(expense_rows)
    export_date  = datetime.now().strftime("%d/%m/%Y %H:%M")
    DATA_START   = 4
    inc_end      = DATA_START + n_inc - 1 if n_inc else DATA_START
    exp_end      = DATA_START + n_exp - 1 if n_exp else DATA_START
    inc_ref      = f"'Income'!C{DATA_START}:C{inc_end}"   if n_inc else "'Income'!C4:C4"
    exp_ref      = f"'Expenses'!C{DATA_START}:C{exp_end}" if n_exp else "'Expenses'!C4:C4"

    # ── SUMMARY ──────────────────────────────────────────────────────────
    ss = wb.active
    ss.title = "Summary"
    ss.sheet_properties.tabColor = NAVY
    ss.sheet_view.showGridLines  = False
    ss.column_dimensions["A"].width = 36
    ss.column_dimensions["B"].width = 24
    ss.column_dimensions["C"].width = 26
    title_cell(ss, "A1:C1", "FINANCIAL SUMMARY", BLUE_DARK, row_h=44)

    ss.row_dimensions[2].height = 8
    for col in range(1, 4):
        ss.cell(row=2, column=col).fill = solid("ECF1FB")

    def section_hdr(row_num, text, bg, fg=WHITE):
        ss.row_dimensions[row_num].height = 20
        for col in range(1, 4):
            c = ss.cell(row=row_num, column=col)
            c.fill   = solid(bg)
            c.border = Border(bottom=_side("thin", bg))
        c0 = ss.cell(row=row_num, column=1)
        c0.value     = text
        c0.font      = Font(name=FONT, bold=True, size=10, color=_argb(fg))
        c0.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c0.border    = Border(left=_side("medium", bg), bottom=_side("thin", bg))

    def s_row(row_num, label, value, context, bg, fg, accent,
              bold_val=False, fmt=TND_FMT, text_val=False):
        ss.row_dimensions[row_num].height = 26
        lc           = ss.cell(row=row_num, column=1, value=label)
        lc.font      = Font(name=FONT, bold=True, size=10, color=_argb(fg))
        lc.fill      = solid(bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        lc.border    = _left_accent(accent)
        vc = ss.cell(row=row_num, column=2)
        if text_val:
            vc.value, vc.number_format = value, "@"
        elif value is None:
            vc.value = ""
        elif isinstance(value, str) and value.startswith("="):
            vc.value, vc.number_format = value, fmt
        else:
            try:   vc.value = float(value)
            except: vc.value = value
            vc.number_format = fmt
        vc.font      = Font(name=FONT, bold=bold_val, size=10, color=_argb(fg))
        vc.fill      = solid(bg)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border    = _box(BORDER_CLR)
        cc           = ss.cell(row=row_num, column=3, value=context)
        cc.font      = Font(name=FONT, size=9, color=_argb(fg))
        cc.fill      = solid(bg)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border    = _right_accent(accent)

    def spacer(row_num, bg="ECF1FB"):
        ss.row_dimensions[row_num].height = 8
        for col in range(1, 4):
            c = ss.cell(row=row_num, column=col)
            c.fill   = solid(bg)
            c.border = Border()

    inc_label = f"{n_inc} {'entry' if n_inc == 1 else 'entries'}"
    exp_label = f"{n_exp} {'entry' if n_exp == 1 else 'entries'}"

    section_hdr(3, "  BALANCE OVERVIEW", BLUE_MID)
    s_row(4,  "Opening Balance", str(opening),        "Starting funds",                                BLUE_LIGHT, BLUE_FG, BLUE_DARK)
    s_row(5,  "Total Income",   f"=SUM({inc_ref})",   inc_label,                                       GREEN_LIGHT,GREEN_FG,GREEN_DARK)
    s_row(6,  "Total Expenses", f"=SUM({exp_ref})",   exp_label,                                       RED_LIGHT,  RED_FG,  RED_DARK)
    spacer(7)
    section_hdr(8, "  RESULTS", BLUE_MID)
    s_row(9,  "Gross Profit",  "=B5-B6",             '=IF(B9>=0,"Surplus \u25b2","Deficit \u25bc")',    BLUE_LIGHT, BLUE_FG, BLUE_DARK, fmt=TND_FMT)
    s_row(10, "Expense Ratio", "=IFERROR(B6/B5,0)",  '=IFERROR(TEXT(B10,"0.0%")&" of income","N/A")',   BLUE_LIGHT, BLUE_FG, BLUE_DARK, fmt=PCT_FMT)
    s_row(11, "Total Profit",  "=B9",                '=IF(B11>=0,"Surplus \u25b2","Deficit \u25bc")',    BLUE_LIGHT, BLUE_FG, BLUE_DARK, fmt=TND_FMT)
    s_row(12, "Net Balance",   "=B4+B11",            '=IF(B12>=0,"Positive \u2714","Negative \u2716")',  BLUE_MID,   WHITE,   BLUE_DARK, bold_val=True, fmt=TND_FMT)
    spacer(13)
    section_hdr(14, "  METADATA", META_ACC)
    s_row(15, "Total Entries",  n_inc + n_exp,        f"total ({n_inc} income, {n_exp} expense)",        META_BG, META_FG, META_ACC, fmt="General")
    s_row(16, "Last Exported",  export_date,          "Export timestamp",                                META_BG, META_FG, META_ACC, text_val=True)

    # ── INCOME ───────────────────────────────────────────────────────────
    ws_inc = wb.create_sheet("Income")
    ws_inc.sheet_properties.tabColor = GREEN_DARK
    ws_inc.sheet_view.showGridLines  = False
    ws_inc.column_dimensions["A"].width =  7
    ws_inc.column_dimensions["B"].width = 36
    ws_inc.column_dimensions["C"].width = 20
    ws_inc.column_dimensions["D"].width = 18
    ws_inc.column_dimensions["E"].width = 44
    ws_inc.row_dimensions[2].height = 6
    ws_inc.row_dimensions[3].height = 24
    ws_inc.freeze_panes = "A4"
    title_cell(ws_inc, "A1:E1", "INCOME", GREEN_DARK, row_h=40)
    for col, (text, align) in enumerate([
        ("ID","center"),("Description","left"),("Amount (TND)","center"),("Date","center"),("Notes","left")
    ], 1):
        hdr(ws_inc.cell(row=3, column=col), text, GREEN_MID, align=align)
    for i, r in enumerate(income_rows):
        row = i + DATA_START
        alt = i % 2 == 1
        ws_inc.row_dimensions[row].height = 22
        dat(ws_inc.cell(row=row, column=1), i+1,              align="center", alt=alt)
        dat(ws_inc.cell(row=row, column=2), r["description"], align="left",   alt=alt)
        dat(ws_inc.cell(row=row, column=3), float(r["amount"]),align="right", alt=alt, fmt=TND_FMT)
        dat(ws_inc.cell(row=row, column=4), r["date"],        align="center", alt=alt)
        dat(ws_inc.cell(row=row, column=5), r.get("notes",""),align="left",   alt=alt, wrap=True)

    # ── EXPENSES ─────────────────────────────────────────────────────────
    ws_exp = wb.create_sheet("Expenses")
    ws_exp.sheet_properties.tabColor = RED_DARK
    ws_exp.sheet_view.showGridLines  = False
    ws_exp.column_dimensions["A"].width =  7
    ws_exp.column_dimensions["B"].width = 36
    ws_exp.column_dimensions["C"].width = 20
    ws_exp.column_dimensions["D"].width = 18
    ws_exp.column_dimensions["E"].width = 44
    ws_exp.row_dimensions[2].height = 6
    ws_exp.row_dimensions[3].height = 24
    ws_exp.freeze_panes = "A4"
    title_cell(ws_exp, "A1:E1", "EXPENSES", RED_DARK, row_h=40)
    for col, (text, align) in enumerate([
        ("ID","center"),("Description","left"),("Amount (TND)","center"),("Date","center"),("Notes","left")
    ], 1):
        hdr(ws_exp.cell(row=3, column=col), text, RED_MID, align=align)
    for i, r in enumerate(expense_rows):
        row = i + DATA_START
        alt = i % 2 == 1
        ws_exp.row_dimensions[row].height = 22
        dat(ws_exp.cell(row=row, column=1), i+1,              align="center", alt=alt)
        dat(ws_exp.cell(row=row, column=2), r["description"], align="left",   alt=alt)
        dat(ws_exp.cell(row=row, column=3), float(r["amount"]),align="right", alt=alt, fmt=TND_FMT)
        dat(ws_exp.cell(row=row, column=4), r["date"],        align="center", alt=alt)
        dat(ws_exp.cell(row=row, column=5), r.get("notes",""),align="left",   alt=alt, wrap=True)

    # ── LOG ──────────────────────────────────────────────────────────────
    ws_log = wb.create_sheet("Log")
    ws_log.sheet_properties.tabColor = CHARCOAL
    ws_log.sheet_view.showGridLines  = False
    ws_log.column_dimensions["A"].width = 26
    ws_log.column_dimensions["B"].width = 14
    ws_log.column_dimensions["C"].width = 12
    ws_log.column_dimensions["D"].width = 40
    ws_log.column_dimensions["E"].width = 24
    ws_log.row_dimensions[2].height = 6
    ws_log.row_dimensions[3].height = 24
    ws_log.freeze_panes = "A4"
    title_cell(ws_log, "A1:E1", "ACTIVITY LOG", CHARCOAL, row_h=40)
    for col, (text, align) in enumerate([
        ("Date","center"),("Action","center"),("Type","center"),("Description","left"),("Amount","center")
    ], 1):
        hdr(ws_log.cell(row=3, column=col), text, GREY_DARK, align=align)

    ACTION_BG = {"ADD":"D4EDDA","REMOVE":"F8D7DA","CLEAR":"FFF3CD","BALANCE":"D6E4F7","IMPORT":"E8D5F5","EXPORT":"D6EAF8"}
    ACTION_FG = {"ADD":"155724","REMOVE":"721C24","CLEAR":"856404","BALANCE":"0D2B5E","IMPORT":"4A0E72","EXPORT":"1A5276"}

    log_rows = _read_log(newest_first=False)
    if not log_rows:
        c = ws_log.cell(row=4, column=1, value="No log entries yet.")
        c.font      = Font(name=FONT, size=10, color=_argb("999999"), italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
    else:
        for i, r in enumerate(log_rows):
            row    = i + DATA_START
            action = r.get("action","").strip().upper()
            abg    = ACTION_BG.get(action, "F5F5F5")
            afg    = ACTION_FG.get(action, "444444")
            ws_log.row_dimensions[row].height = 22
            dat(ws_log.cell(row=row, column=1), r.get("date","").strip(),        align="center")
            ac            = ws_log.cell(row=row, column=2, value=action)
            ac.font       = Font(name=FONT, bold=True, size=9, color=_argb(afg))
            ac.fill       = solid(abg)
            ac.alignment  = Alignment(horizontal="center", vertical="center")
            ac.border     = _box(BORDER_CLR)
            dat(ws_log.cell(row=row, column=3), r.get("type","").strip(),        align="center")
            dat(ws_log.cell(row=row, column=4), r.get("description","").strip(), align="left")
            amt = r.get("amount","").strip()
            try:    dat(ws_log.cell(row=row, column=5), float(amt), align="right", fmt=TND_FMT)
            except: dat(ws_log.cell(row=row, column=5), amt,        align="right")

    # ── Save xlsx ─────────────────────────────────────────────────────────
    try:
        wb.save(EXPORT_FILE)
    except OSError as e:
        print(f"Error: could not save '{EXPORT_FILE}': {e}", file=sys.stderr)
        sys.exit(EXIT_FILE_ERROR)

    print(f"Exported '{EXPORT_FILE}' — Summary | Income ({n_inc}) | Expenses ({n_exp}) | Log ({len(log_rows)}).")

    # ══════════════════════════════════════════════════════════════════════
    # GOOGLE SHEETS PUSH
    # ══════════════════════════════════════════════════════════════════════
    if not (push_to_gsheet and gsheet_id):
        return

    print(f"\nPushing to Google Sheets ({gsheet_id})...")

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build

        creds   = service_account.Credentials.from_service_account_file(
            GSHEET_CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        service = build("sheets", "v4", credentials=creds)

        total_inc = sum(float(r["amount"]) for r in income_rows)
        total_exp = sum(float(r["amount"]) for r in expense_rows)
        gross     = total_inc - total_exp
        net       = opening + gross
        exp_pct   = f"{(total_exp / total_inc * 100):.1f}%" if total_inc else "N/A"

        summary_data = [
            ["FINANCIAL SUMMARY", "", ""],
            ["", "", ""],
            ["BALANCE OVERVIEW", "", ""],
            ["Opening Balance", opening,    "Starting funds"],
            ["Total Income",    total_inc,  f"{n_inc} {'entry' if n_inc == 1 else 'entries'}"],
            ["Total Expenses",  total_exp,  f"{n_exp} {'entry' if n_exp == 1 else 'entries'}"],
            ["", "", ""],
            ["RESULTS", "", ""],
            ["Gross Profit",  gross,   "Surplus \u25b2" if gross >= 0 else "Deficit \u25bc"],
            ["Expense Ratio", exp_pct, "of income"],
            ["Total Profit",  gross,   "Surplus \u25b2" if gross >= 0 else "Deficit \u25bc"],
            ["Net Balance",   net,     "Positive \u2714" if net >= 0 else "Negative \u2716"],
            ["", "", ""],
            ["METADATA", "", ""],
            ["Total Entries", n_inc + n_exp, f"total ({n_inc} income, {n_exp} expense)"],
            ["Last Exported", export_date,   ""],
        ]
        # Each data array must include title (row 1) + spacer (row 2) + headers (row 3)
        # so the values write aligns with the batchUpdate format requests.
        income_data  = [
            ["INCOME", "", "", "", ""],
            ["", "", "", "", ""],
            ["ID", "Description", "Amount (TND)", "Date", "Notes"],
        ]
        for i, r in enumerate(income_rows, 1):
            income_data.append([i, r["description"], float(r["amount"]), r["date"], r.get("notes","")])

        expense_data = [
            ["EXPENSES", "", "", "", ""],
            ["", "", "", "", ""],
            ["ID", "Description", "Amount (TND)", "Date", "Notes"],
        ]
        for i, r in enumerate(expense_rows, 1):
            expense_data.append([i, r["description"], float(r["amount"]), r["date"], r.get("notes","")])

        log_data = [
            ["ACTIVITY LOG", "", "", "", ""],
            ["", "", "", "", ""],
            ["Date", "Action", "Type", "Description", "Amount"],
        ]
        for r in log_rows:
            amt = r.get("amount","").strip()
            try:    amt = float(amt)
            except: pass
            log_data.append([r.get("date","").strip(), r.get("action","").strip(),
                              r.get("type","").strip(), r.get("description","").strip(), amt])

        tabs_data = {"Summary": summary_data, "Income": income_data,
                     "Expenses": expense_data, "Log": log_data}

        # Delete + recreate tabs cleanly
        spreadsheet  = service.spreadsheets().get(spreadsheetId=gsheet_id).execute()
        existing_ids = {s["properties"]["title"]: s["properties"]["sheetId"]
                        for s in spreadsheet.get("sheets", [])}

        service.spreadsheets().batchUpdate(
            spreadsheetId=gsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": "__tmp__"}}}]},
        ).execute()

        delete_reqs = [{"deleteSheet": {"sheetId": sid}}
                       for title, sid in existing_ids.items() if title in tabs_data]
        if delete_reqs:
            service.spreadsheets().batchUpdate(
                spreadsheetId=gsheet_id, body={"requests": delete_reqs}).execute()

        service.spreadsheets().batchUpdate(
            spreadsheetId=gsheet_id,
            body={"requests": [
                {"addSheet": {"properties": {"title": t, "index": i}}}
                for i, t in enumerate(tabs_data)
            ]},
        ).execute()

        refreshed = service.spreadsheets().get(spreadsheetId=gsheet_id).execute()
        tmp_id    = next(s["properties"]["sheetId"] for s in refreshed["sheets"]
                         if s["properties"]["title"] == "__tmp__")
        service.spreadsheets().batchUpdate(
            spreadsheetId=gsheet_id,
            body={"requests": [{"deleteSheet": {"sheetId": tmp_id}}]},
        ).execute()

        # Collect new tab IDs
        final      = service.spreadsheets().get(spreadsheetId=gsheet_id).execute()
        tab_ids    = {s["properties"]["title"]: s["properties"]["sheetId"]
                      for s in final["sheets"] if s["properties"]["title"] in tabs_data}

        # Write data
        for tab_name, data in tabs_data.items():
            result = service.spreadsheets().values().update(
                spreadsheetId=gsheet_id,
                range=f"'{tab_name}'!A1",
                valueInputOption="USER_ENTERED",
                body={"values": data},
            ).execute()
            print(f"  \u2713 {tab_name}: {result.get('updatedCells', 0)} cells written")

        # Apply formatting via batchUpdate
        print("  Applying formatting...")
        fmt_requests = _build_gsheet_format_requests(tab_ids, n_inc, n_exp, log_rows)
        # Send in chunks of 500 to stay well within API limits
        chunk = 500
        for i in range(0, len(fmt_requests), chunk):
            service.spreadsheets().batchUpdate(
                spreadsheetId=gsheet_id,
                body={"requests": fmt_requests[i:i+chunk]},
            ).execute()
        print(f"  \u2713 Formatting applied ({len(fmt_requests)} requests)")

        log_event("EXPORT", "", f"Pushed to Google Sheets {gsheet_id}", "")
        print(f"\n\u2713 Successfully pushed to Google Sheets.")

    except Exception as e:
        print(f"Error pushing to Google Sheets: {e}", file=sys.stderr)
        print("Local export completed; Google Sheets push failed.", file=sys.stderr)


main()